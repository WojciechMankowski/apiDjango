from openpyxl import load_workbook

def load_data_from_excel(file_path, sheet_name):
    skoroszyt = load_workbook(file_path)
    if sheet_name in skoroszyt.sheetnames:
        arkusz = skoroszyt[sheet_name]
    else:
        print(f"Arkusz '{sheet_name}' nie istnieje w pliku.")
        return []

    # Pobranie nagłówków z pierwszego wiersza
    naglowki = [cell for cell in next(arkusz.iter_rows(values_only=True))]
    dane = []

    # Przechodzenie przez każdy wiersz po pierwszym, tworząc słowniki
    for wiersz in arkusz.iter_rows(min_row=2, values_only=True):
        rekord = dict(zip(naglowki, wiersz))
        dane.append(rekord)

    return dane

if __name__ == '__main__':
    sciezka_pliku = 'dane.xlsx'
    nazwa_arkusza = 'Dane'
    dane_z_excela = load_data_from_excel(sciezka_pliku, nazwa_arkusza)
    print(dane_z_excela)
